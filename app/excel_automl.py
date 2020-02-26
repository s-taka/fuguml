# -*- coding: utf-8 -*-

import os, sys
parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, parent_dir)

from fuguml.automl import AutoMLTable
from fuguml.automl import CONSTANT

import pprint as pp
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart
from openpyxl.chart import Reference

import logging
import re
import datetime

# for Pyinstaller
from scipy import sparse
import lightgbm
import optuna
import sklearn

logging.basicConfig(format='%(asctime)s : %(levelname)s : %(message)s', level=logging.INFO)
logger = logging.getLogger()
logger.setLevel(0)

def write_excel_sheet(wb_out, sheet_name, df, min_len = -1, index=True, header=True):
    ws = wb_out.create_sheet(title=sheet_name)
    for r in dataframe_to_rows(df, index=index, header=header):
        if len(r) > min_len:
            ws.append(r)
    return ws

def main():
    if len(sys.argv) != 2:
        print("Usage: $ python %s excel_file" % sys.argv[0])
        quit()

    excel_file = sys.argv[1]
    logger.info(excel_file)

    wb = load_workbook(excel_file, read_only=True)
    out_excel_file = "{}_result.xlsx".format(re.sub(r"\.[^\.]+$", "", excel_file))
    wb_out = Workbook()
    ws_summary = wb_out.active
    ws_summary.title = "summary"
    ws_summary['A1'] = "FuguAutoML summary file"
    ws_summary['A2'] = "create timedate"
    ws_summary['B2'] = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    wb_out.save(out_excel_file)

    logger.info("out_excel_file = {}".format(out_excel_file))

    logger.info("load conf sheet")
    conf_sheet = wb["conf"]
    conf = {}
    for i in range(1, conf_sheet.max_row + 1):
        key_cel = conf_sheet[i][0].value
        val_cel = conf_sheet[i][2].value
        if key_cel:
            conf[key_cel] = val_cel

    logger.info(pp.pformat(conf))

    table_structure_sheet = conf["table_structure"]
    main_table = conf["main_table"]
    operation = conf["operation"]
    objective_column = conf["objective_column"]
    key_column = conf["key_column"]
    apply_table = conf["apply_table"]

    logger.info("load table_structure sheet")
    table_sheet = wb[table_structure_sheet]
    table_structure = {}
    for i in range(2, table_sheet.max_row + 1):
        table_name = table_sheet[i][0].value
        col_name = table_sheet[i][1].value
        ope = table_sheet[i][2].value
        if table_name and col_name and ope:
            if table_name not in table_structure:
                table_structure[table_name] = {}
            if col_name not in table_structure[table_name]:
                table_structure[table_name][col_name] = ope
    logger.info(pp.pformat(table_structure))

    logger.info("load data")
    Xs = {}
    X_apply = None
    for data_table in table_structure:
        data_sheet = wb[data_table]
        data_type = {}
        use_col = []
        for i in range(data_sheet.max_column):
            col_name = data_sheet[1][i].value
            if col_name in table_structure[data_table]:
                use_col.append(i)
                ope = table_structure[data_table][col_name]
                if ope == CONSTANT.KEY_TYPE or ope == CONSTANT.CATEGORY_TYPE or \
                        ope == CONSTANT.TEXT_TYPE or ope == CONSTANT.TEXTJA_TYPE:
                    data_type[col_name] = str
                elif ope == CONSTANT.NUMERICAL_TYPE:
                    data_type[col_name] = np.float64
                elif ope == CONSTANT.OBJ_TYPE:
                    data_type[col_name] = np.int64
                else:
                    logger.warn("type handling error")

        pd_sheet = pd.read_excel(excel_file, sheet_name=data_table, usecols=use_col,
                                 dtype=data_type)
        Xs[data_table] = pd_sheet
        if data_table == main_table:
            X_apply = pd.read_excel(excel_file, sheet_name=apply_table, usecols=use_col,
                                 dtype=data_type)
    logger.info(pp.pformat(Xs))

    for ws_name in Xs.keys():
        write_excel_sheet(wb_out, "{}_desc".format(ws_name), Xs[ws_name].describe(include="all"), min_len=1)
    write_excel_sheet(wb_out, "{}_desc".format(apply_table), X_apply.describe(include="all"), min_len=1)

    logger.info("new AutoML instance")
    amt_struct = {}
    for t in table_structure:
        for c in table_structure[t]:
            if t not in amt_struct:
                amt_struct[t] = []
            if table_structure[t][c] != CONSTANT.OBJ_TYPE:
                amt_struct[t].append((c, table_structure[t][c]))
    logger.info(amt_struct)

    amt = AutoMLTable({"struct": amt_struct}, key_col=key_column, main_tbl=main_table, apply_tbl=apply_table)
    amt.n_trials = 100
    obj = np.array(Xs[main_table][objective_column]).ravel()
    scores, importance, proba_importances, cv_result, proba_cv_result, main_features = amt.make_model(Xs, obj)
    pp.pprint(proba_importances)

    ws_summary["A3"] = "cv scores (roc auc)"
    for i, s in enumerate(scores):
        ws_summary.cell(column=2+i, row=3, value=s)
    ws_summary["A4"] = "cv scores prob (roc auc)"
    for i, s in enumerate(proba_cv_result["scores"]):
        ws_summary.cell(column=2+i, row=4, value=s)

    imp_data = importance[0]
    importance_sheet = pd.DataFrame()
    importance_sheet["feature_name"] = [im[2] for im in imp_data]
    importance_sheet["all_gain"] = [im[0] for im in imp_data]
    importance_sheet["all_split"] = [im[1] for im in imp_data]
    for i, imp in enumerate(proba_importances):
        importance_sheet["all_probability-{}_gain".format(i)] = [imp[im[2]][0] for im in importance[0]]
    for i, imp in enumerate(cv_result["importances"]):
        importance_sheet["CVfold-{}_gain".format(i)] = [imp[1][im[2]][0] for im in importance[0]]
        importance_sheet["CVfold-{}_split".format(i)] = [imp[1][im[2]][1] for im in importance[0]]
    for j, imps in enumerate(proba_cv_result["importances"]):
        for i, imp in enumerate(imps):
            importance_sheet["CVfold_probability-{}-{}_gain".format(j, i)] = [imp[im[2]][0] for im in importance[0]]
    write_excel_sheet(wb_out, "importance", importance_sheet, index=False)

    cv_sheet = pd.DataFrame()
    cv_sheet["predict_score"] = cv_result["preds"]
    cv_sheet["fold_num"] = cv_result["preds_fold_num"]
    cv_sheet["calibrated_probability"] = proba_cv_result["preds"]
    cv_sheet["rank_in_fold"] = [-1 for i in cv_result["preds_fold_num"]]
    cv_sheet["rank_percentile"] = [-1 for i in cv_result["preds_fold_num"]]
    for fold_num, m in enumerate(cv_result["models"]):
        cv_sheet.loc[cv_sheet["fold_num"] == fold_num, "rank_in_fold"] = \
            cv_sheet[cv_sheet["fold_num"] == fold_num]["predict_score"].rank()
        cv_sheet.loc[cv_sheet["fold_num"] == fold_num, "rank_percentile"] = \
            cv_sheet[cv_sheet["fold_num"] == fold_num]["rank_in_fold"] / \
            cv_sheet[cv_sheet["fold_num"] == fold_num]["rank_in_fold"].max()
    cv_sheet[key_column] = Xs[main_table][key_column]
    cv_sheet["truth"] = obj
    for mf in main_features:
        cv_sheet["feature[{}]".format(mf[0])] = mf[1].todense()
    write_excel_sheet(wb_out, "cv_result", cv_sheet, index=False)

    for_gain_graph = pd.DataFrame()
    for_gain_graph["predict_score"] = [f for f in cv_sheet["predict_score"]]
    for_gain_graph["rank"] = [f for f in cv_sheet["rank_percentile"]]
    for_gain_graph["truth"] = [float(t) for t in cv_sheet["truth"]]
    for_gain_graph[key_column] = [k for k in cv_sheet[key_column]]
    for_gain_graph["fold_num"] = [k for k in cv_sheet["fold_num"]]
    sum_truth = for_gain_graph["truth"].sum()
    subsum_truth = 0.0
    perfect_truth = 0.0
    gain_list = []
    perfect_list = []
    random_list = []
    for_gain_graph.sort_values("rank", inplace=True, ascending=False)
    for idx, row in enumerate(for_gain_graph.itertuples()):
        obj = row[3]
        subsum_truth += obj
        perfect_truth += 1.0
        gain_list.append(subsum_truth / sum_truth)
        perfect_list.append(min(perfect_truth / sum_truth, 1.0))
        random_list.append((idx + 1.0) / len(for_gain_graph["rank"]))
    for_gain_graph["this_model"] = gain_list
    for_gain_graph["perfect_model"] = perfect_list
    for_gain_graph["random"] = random_list
    for_gain_graph["threasolds"] = ["=HLOOKUP(E{}, summary!A6:Z7, 2, FALSE)".format(2+i)
                                    for i, r in enumerate(random_list)]
    for_gain_graph["predict_label"] = ["=IF(A{} >= I{}, 1, 0)".format(2+i, 2+i) for i, r in enumerate(random_list)]
    for_gain_graph["flag"] = ["=INT(100 + C{}*10 +  J{})".format(2+i, 2+i) for i, r in enumerate(random_list)]
    ws = write_excel_sheet(wb_out, "gain", for_gain_graph, index=False)

    c1 = LineChart()
    c1.title = "Gain chart"
    c1.y_axis.title = 'cumulative gain'
    c1.x_axis.title = 'rank'

    data = Reference(ws, min_col=6, min_row=1, max_col=8, max_row=1 + len(gain_list))
    c1.add_data(data, titles_from_data=True)
    s2 = c1.series[1]
    s2.graphicalProperties.line.dashStyle = "sysDot"
    s3 = c1.series[2]
    s3.graphicalProperties.line.dashStyle = "sysDot"
    ws_summary["A20"] = "gain chart"
    ws_summary.add_chart(c1, "A21")
    ws_summary["A6"] = "model no"
    for i, s in enumerate(range(len(cv_result["threasholds"]))):
        ws_summary.cell(column=2+i, row=6, value=s)
    ws_summary["A7"] = "threashold"
    for i, s in enumerate(cv_result["threasholds"]):
        ws_summary.cell(column=2+i, row=7, value=s)
    ws_summary["A8"] = "precision"
    ws_summary["A9"] = "recall"
    ws_summary["A10"] = "f-value"
    ws_summary["C11"] = "truth"
    ws_summary["A13"] = "predict"
    ws_summary["C12"] = "1"
    ws_summary["D12"] = "0"
    ws_summary["B13"] = "1"
    ws_summary["B14"] = "0"
    ws_summary["C13"] = "=COUNTIF(gain!K:K, 111)"
    ws_summary["D13"] = "=COUNTIF(gain!K:K, 101)"
    ws_summary["C14"] = "=COUNTIF(gain!K:K, 110)"
    ws_summary["D14"] = "=COUNTIF(gain!K:K, 100)"
    ws_summary["B8"] = "=C13/(C13+D13)"
    ws_summary["B9"] = "=C13/(C13+C14)"
    ws_summary["B10"] = "=(2*B9*B8)/(B8+B9)"

    ws_summary["A16"] = "TOP"
    ws_summary["B16"] = 0.1
    ws_summary['B16'].number_format = '0.0%'
    ws_summary["C16"] = "capture"
    ws_summary["D16"] = '=SUM(INDIRECT("gain!C2:C"&INT(SUM(C13:D14)*(B16))+1))/SUM(INDIRECT("gain!C2:C"&INT(SUM(C13:D14))+1))'
    ws_summary['D16'].number_format = '0.0%'



    Xs[apply_table] = X_apply
    del Xs[main_table]
    pred, pred_proba, main_features = amt.apply(Xs, proba=True, out_main_feature=True)
    pred_sheet = pd.DataFrame()
    pred_sheet["predict_score"] = pred
    pred_sheet["calibrated_probability"] = pred_proba
    pred_sheet[key_column] = X_apply[key_column]
    for mf in main_features:
        pred_sheet["feature[{}]".format(mf[0])] = mf[1].todense()
    write_excel_sheet(wb_out, "predict_all", pred_sheet, index=False)

    pred, pred_proba, main_features = amt.apply_cv(Xs, proba=True, out_main_feature=True)
    pred_sheet = pd.DataFrame()
    pred_sheet["predict_score"] = pred
    pred_sheet["calibrated_probability"] = pred_proba
    pred_sheet[key_column] = X_apply[key_column]
    for mf in main_features:
        pred_sheet["feature[{}]".format(mf[0])] = mf[1].todense()
    write_excel_sheet(wb_out, "predict_cv", pred_sheet, index=False)

    wb_out.save(out_excel_file)
    return


if __name__ == "__main__":
    main()